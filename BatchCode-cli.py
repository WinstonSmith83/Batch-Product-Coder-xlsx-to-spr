from os.path import exists as file_exists
import time
import sys
from pathlib import Path
import os
from rich import print
from openpyxl import load_workbook
from time import sleep
from random import randint
from rich.console import Console
from rich.text import Text
from rich.panel import Panel
from tkinter import Tk  
from tkinter.filedialog import askopenfilename, askdirectory
#etime = int(time.time())
#if etime > 1665629979:
#    sys.exit("Limited Test Over")
#    sleep(10)
#tkinter console Panel with welcome message
console = Console()
Tk().withdraw() 
welcomemsg = Panel("[bold grey89]Please enter columns to pull in this order: Lot, BB, Product, UPC.\nExample A C G R, will pull all values from the A, C, G and R columns,\nin that exact order. Each entry must be sperated by a space. ", title="Batch Encoder File Generator", expand="False")
console.print(welcomemsg)
#get input for which columns to pull data from
lotcol, bbcol, prodcol, barcol = input("Please Enter Columns to pull data from:").split() 
#tk dialog ask for excel file then ask what directory to save to  
filename = askopenfilename(title="Select Microsoft Excel File")
directory = askdirectory(title="Please select USB Drive to save files to.")
#start time of program
startTime = time.time()
#load excel file
workbook = load_workbook(filename=filename)
workbook.iso_dates = True
workbook.sheetnames
sheet = workbook.active
#set variables for later 
lot = None
bb = None
prod = None
bar = None
col1 = []
col2 = []
col3 = []
col4 = []
#for loop read lot, bb, prod, bar from .xlsx file into list
for i in sheet[lotcol]:
    col1.append(i.value)
for i in sheet[bbcol]:
    col2.append(i.value)
for i in sheet[prodcol]:
    col3.append(i.value)
for i in sheet[barcol]:
    col4.append(i.value)
for col01, col02, col03, col04 in zip(col1, col2, col3, col4):
        lot = str(col01)
        bb = col02.strftime("%x")
        prod = col03
        bar = col04
        file = f"""{{
    "BarcodeItemInfo": [
        {{
            "BarAtX": 22,
            "BarAtY": 303,
            "BarContent": \"{bar}\",
            "BarIsExtVariable": 0,
            "BarIsTextOn": false,
            "BarLineHeight": 33,
            "BarLineWidth": 8,
            "BarMode": 20,
            "BarRotate": 0
        }}
    ],
    "SceneWidth": 1741,
    "TextItemInfo": [
        {{
            "TextAtX": 23,
            "TextAtY": 52,
            "TextContent": \"{lot}\",
            "TextFont": "simsun,-1,150,5,50,0,0,0,0,0",
            "TextIsExtVariable": 0,
            "TextRotate": 0,
            "TextSpace": 0
        }},
        {{
            "TextAtX": 892,
            "TextAtY": 299,
            "TextContent": \"{bb}\",
            "TextFont": "simsun,-1,153,5,50,0,0,0,0,0",
            "TextIsExtVariable": 0,
            "TextRotate": 0,
            "TextSpace": 0
        }},
        {{
            "TextAtX": 895,
            "TextAtY": 48,
            "TextContent": \"{prod}\",
            "TextFont": "simsun,-1,153,5,50,0,0,0,0,0",
            "TextIsExtVariable": 0,
            "TextRotate": 0,
            "TextSpace": -7
        }}
    ]
}}
        """
        #Set path to file to be written
        here = directory #os.path.dirname(os.path.abspath(__file__))
        path_to_file = os.path.join(directory, lot+"-"+prod+".spr")
        path_to_file_fancy = Panel("[grey89]File Written:  " + path_to_file, title="Working...", expand="False")
        path = Path(path_to_file)
        #Check if file already exists before writing
        if path.is_file():
            allpretty = Panel("[bold dark_red] Skipping [bold grey89] " + path_to_file + "[bold dark_red] Already Exists", title="Error", expand="False")
            console.print(allpretty)
            continue  
        f = open(path, "a")
        f.write(file)
        f.close() 
        if path.is_file():
            console.print(path_to_file_fancy)         
            #Random sleep, just to make the output look better.
            sleep(randint(1,2))
        else:
            #if we hit this we are fucked, need to put a hard exit here.
            sys.exit("Danger!! Will Robinson, Danger!!\n Unknown Error, Exiting Program")                   

executionTime = (time.time() - startTime)
complete = Text("Finished parsing .xlsx file and writing .spr layout files")
complete.stylize("bold deep_pink4")
console.print(complete)
console.print('Execution time in seconds: ' + str(executionTime))
x = input("Press any key to exit")


